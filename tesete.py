import openpyxl

# criar uma planilha(book)
book = openpyxl.Workbook()

# como vizualiuzar paginas existentes
print(book.sheetnames)

#como criar uma pagina
book.create_sheet('frutas')

#como selecionar uma pagina
sheet = book['frutas']
sheet.append(['frutas', 'quantidade', 'valor'])
sheet.append(['banana', '5', 'R$5,00'])
sheet.append(['frutas1', '5', 'R$3,00'])
sheet.append(['frutas2', '5', 'R$5,00'])
sheet.append(['frutas3', '5', 'R$4,00'])

#salvar a planilha com o nome
book.save('frutas.xlsx')